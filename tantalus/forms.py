import os
import datetime
import json
import pandas as pd

#===========================
# Django imports
#---------------------------
from django import forms

#===========================
# App imports
#---------------------------
from django.db import transaction
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError

import tantalus.models

from openpyxl import load_workbook


class AnalysisForm(forms.ModelForm):
    class Meta:
        model = tantalus.models.Analysis
        fields = [
            'name',
            'analysis_type',
            'version',
            'jira_ticket',
            'args',
        ]


class PatientForm(forms.ModelForm):
    class Meta:
        model = tantalus.models.Patient
        fields = '__all__'
        help_texts = {
            'patient_id': 'Next available SA ID. You may change this if you wish',
        }

    def clean_patient_id(self):
        patient_id = self.cleaned_data.get('patient_id', False)
        if(patient_id[:2] != "SA"):
            raise ValidationError("Error!. Patient IDs must start with SA")
        return patient_id


class UploadPatientForm(forms.Form):
    patients_excel_file = forms.FileField(label='Select an Excel File', widget=forms.FileInput(attrs={'accept': [".xlsx", ".xls"]}))

    def clean_patients_excel_file(self):
        excel_file = self.cleaned_data.get("patients_excel_file", False)
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        temp_dict = {}
        sheet_rows = list(sheet.rows)
        header_row = []
        for cell in sheet_rows[0]:
            header_row.append(cell.value.encode('utf-8'))

        try:
            external_patient_id_index = header_row.index('External Patient ID')
            patient_id_index = header_row.index('Patient ID')
            case_id_index = header_row.index('Case ID')
        except:
            raise ValidationError('Header Row Labels not configured properly')

        sheet_rows.pop(0)

        for index in range(0, len(header_row)):
            column_data_list = []
            for cell in list(sheet.columns)[index]:
                column_data_list.append(cell.value)
            #Remove header column entry from column_data_list
            column_data_list.pop(0)
            temp_dict[header_row[index]] = column_data_list

        df = pd.DataFrame(data=temp_dict)

        SA_prefix_patients = tantalus.models.Patient.objects.filter(patient_id__startswith='SA').order_by('-patient_id')
        patient_ids = []
        
        for patient in SA_prefix_patients:
            patient_ids.append(int(patient.patient_id[2:]))
        patient_ids.sort()

        form_headers = df.columns.tolist()

        external_patient_id_index = form_headers.index('External Patient ID')
        case_id_index = form_headers.index('Case ID')
        patient_id_index = form_headers.index('Patient ID')

        next_available_patient_id = patient_ids[-1] + 1
        auto_generated_patient_ids = []

        for idx, patient_row in df.iterrows():
            if(pd.isnull(patient_row[case_id_index])):
                raise ValidationError("Error on Row {}. Case ID cannot be empty".format(idx + 2))
            if(pd.isnull(patient_row[patient_id_index]) and pd.isnull(patient_row[external_patient_id_index])):
                raise ValidationError("Error on Row {}. Both Patient ID and External Patient IDs cannot be empty".format(idx + 2))
            elif(pd.isnull(patient_row[patient_id_index])):
                patient_row[patient_id_index] = 'SA' + str(next_available_patient_id)
                auto_generated_patient_ids.append('SA' + str(next_available_patient_id))
                next_available_patient_id +=1
            elif(pd.isnull(patient_row[external_patient_id_index])):
                raise ValidationError("Error on Row {}. External Patient ID cannot be empty".format(idx + 2))
            if(patient_row[patient_id_index][:2] != "SA"):
                raise ValidationError("Error on Row {}. Patient IDs must start with SA and not be {}".format(idx + 2, patient_row[patient_id_index]))

        return df, auto_generated_patient_ids

    def get_patient_data(self):
        return self.cleaned_data['patients_excel_file']


class SubmissionForm(forms.ModelForm):
    class Meta:
        model = tantalus.models.Submission
        fields = [
            'sample',
            'sow',
            'submission_date',
            'submitted_by',
            'library_type',
            'coverage',
        ]


#===========================
# Sample forms
#---------------------------
class SampleForm(forms.ModelForm):
    class Meta:
        model = tantalus.models.Sample
        fields = [
            'sample_id',
            'external_sample_id',
            'submitter',
            'collaborator',
            'tissue',
            'note',
            'patient_id',
            'projects',
        ]


class UploadSampleForm(forms.Form):
    samples_excel_file = forms.FileField(label='Select an Excel File', widget=forms.FileInput(attrs={'accept': [".xlsx", ".xls"]}))

    def clean_samples_excel_file(self):
        excel_file = self.cleaned_data.get("samples_excel_file", False)
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        temp_dict = {}
        sheet_rows = list(sheet.rows)
        header_row = []
        for cell in sheet_rows[0]:
            header_row.append(cell.value.encode('utf-8'))

        try:
            external_patient_id_index = header_row.index('External Patient ID')
            external_sample_id_index = header_row.index('External Sample ID')
            patient_id_index = header_row.index('Patient ID')
            suffix_index = header_row.index('Suffix')
        except:
            raise ValidationError('Header Row Labels not configured properly')

        sheet_rows.pop(0)

        for index in range(0, len(header_row)):
            column_data_list = []
            for cell in list(sheet.columns)[index]:
                column_data_list.append(cell.value)
            #Remove header column entry from column_data_list
            column_data_list.pop(0)
            temp_dict[header_row[index]] = column_data_list

        df = pd.DataFrame(data=temp_dict)

        for idx, sample_row in df.iterrows():
            if(pd.isnull(sample_row[external_sample_id_index])):
                raise ValidationError('External Sample ID field cannot be Null on row {}'.format(idx + 2))
            #If both fields are null, raise error
            if(pd.isnull(sample_row[external_patient_id_index]) and pd.isnull(sample_row[patient_id_index])):
                raise ValidationError('Both External Patient ID and Patient ID field cannot be Null on row {}'.format(idx + 2))
            #If only external_patient_id is null, look up patient with patient_id
            elif(pd.isnull(sample_row[external_patient_id_index])):
                try:
                    patient = tantalus.models.Patient.objects.get(patient_id=sample_row[patient_id_index])
                except:
                    raise ValidationError('Patient not found with given Patient ID on row {} (No External Patient ID provided)'.format(
                        sample_row[patient_id_index],
                        idx + 2))
                #If given suffix and patient_id match with an existing sample in tantalus, throw error
                if(sample_row[suffix_index] is None):
                    new_sample_id = patient.patient_id
                else:
                    new_sample_id = sample_row[patient_id_index] + sample_row[suffix_index]
                try:
                    sample = tantalus.models.Sample.objects.get(sample_id=new_sample_id)
                except:
                    continue
                raise ValidationError('Sample already exists for Patient ID {} and Suffix {} on row {}'.format(
                    sample_row[patient_id_index],
                    sample_row[suffix_index],
                    idx + 2
                    ))

            elif(pd.isnull(sample_row[patient_id_index])):
                try:
                    patient = tantalus.models.Patient.objects.get(external_patient_id=sample_row[external_patient_id_index])
                except:
                    raise ValidationError('Patient not found with given External Patient ID on row {} (No Patient ID provided)'.format(
                        sample_row[external_patient_id_index],
                        idx + 2))

                if(sample_row[suffix_index] is None):
                    new_sample_id = patient.patient_id
                else:
                    new_sample_id = patient.patient_id + sample_row[suffix_index]
                try:
                    sample = tantalus.models.Sample.objects.get(sample_id=new_sample_id)
                except:
                    continue
                raise ValidationError('Sample already exists for Patient ID {} and Suffix {} on row {}'.format(
                    patient.patient_id,
                    sample_row[suffix_index],
                    idx + 2
                    ))

                    
            else:
                try:
                    patient = tantalus.models.Patient.objects.get(external_patient_id=sample_row[external_patient_id_index], patient_id=sample_row[patient_id_index])
                except:
                    raise ValidationError('Patient not found with given External Patient ID {} and Patient ID {} on row {}'.format(
                        sample_row[external_patient_id_index],
                        sample_row[patient_id_index],
                        idx + 2))         

                if(sample_row[suffix_index] is None):
                    new_sample_id = patient.patient_id
                else:
                    new_sample_id = sample_row[patient_id_index] + sample_row[suffix_index]
                try:
                    sample = tantalus.models.Sample.objects.get(sample_id=new_sample_id)
                except:
                    continue
                raise ValidationError('Sample already exists for Patient ID {} and Suffix {} on row {}'.format(
                    sample_row[patient_id_index],
                    sample_row[suffix_index],
                    idx + 2
                    ))

        return df

    def get_sample_data(self):
        # TODO: return dataframe here
        return self.cleaned_data['samples_excel_file']


class ExternalIDSearcherForm(forms.Form):
    external_id_column = forms.CharField(widget=forms.Textarea, label='Paste in an Excel Column of External Sample IDs')

    def clean_file(self):
        clean_data = self.cleaned_data.get(external_id_column)
        external_id_list = clean_data.split('\n')
        return external_id_list


class ExternalIDSearcherForm(forms.Form):
    external_id_column = forms.CharField(widget=forms.Textarea, label='Paste in an Excel Column of External Sample IDs')

    def clean_file(self):
        clean_data = self.cleaned_data.get(external_id_column)
        external_id_list = clean_data.split('\n')
        return external_id_list


class DatasetSearchForm(forms.Form):
    tagged_with = forms.CharField(
        label="Tagged with",
        help_text="A comma separated list of tags",
        required=False,
    )
    exclude = forms.CharField(
        label="Exclude",
        help_text="A comma separated list of tags you want to exclude",
        required=False,
    )
    library = forms.CharField(
        label="Library",
        required=False,
        help_text="A white space separated list of library IDs. Eg. MF1606301",
        widget=forms.widgets.Textarea
    )
    sample = forms.CharField(
        label="Sample(s)",
        required=False,
        help_text="A white space separated list of sample IDs. Eg. SA928",
        widget=forms.widgets.Textarea
    )

    dataset_type = forms.MultipleChoiceField(
        choices=tantalus.models.SequenceDataset.dataset_type_choices,
        label="Dataset type",
        required=False,
        help_text="Type of files to process",
        widget=forms.widgets.CheckboxSelectMultiple()
    )
    storages = forms.ModelMultipleChoiceField(
        queryset=tantalus.models.Storage.objects.all(),
        required=False,
        help_text="Only look for files that are present in the selected storage.",
        widget=forms.widgets.CheckboxSelectMultiple(),
    )
    compression_schemes = forms.MultipleChoiceField(
        choices=tantalus.models.FileResource.compression_choices,
        required=False,
        help_text="Only look for files with given compression schemes.",
        widget=forms.widgets.CheckboxSelectMultiple(),
    )
    flowcell_id_and_lane = forms.CharField(
        label="Flowcell ID + lane number",
        required=False,
        help_text="A white space separated list of flowcell IDs and lane number. Eg. H3LGYCCXY_4 - H3LGYCCXY is the lane, 4 is the lane number",
        widget = forms.widgets.Textarea
    )

    sequencing_center = forms.ModelChoiceField(
        queryset=tantalus.models.SequencingLane.objects.all().values_list('sequencing_centre').distinct(),
        empty_label='---',
        label="Sequencing center",
        required=False,
        help_text="Sequencing center that the data was obtained from"
    )

    sequencing_instrument = forms.ModelChoiceField(
        queryset=tantalus.models.SequencingLane.objects.all().values_list('sequencing_instrument').distinct(),
        empty_label='---',
        label="Sequencing instrument",
        required=False,
    )

    sequencing_library_id = forms.CharField(
        label="Sequencing library ID",
        required=False,
        help_text="A white space separated list of external sequencing library ids. " + \
                  "Note that this is different from internal library IDs. " + \
                  "For example, these are external library IDs given to us by the GSC, eg. PX0827",
        widget=forms.widgets.Textarea
    )

    library_type = forms.ModelChoiceField(
        queryset=tantalus.models.LibraryType.objects.all(),
        empty_label='---',
        label="Library type",
        required=False,
    )

    index_format = forms.ChoiceField(
        choices=(('', '---'),) + tantalus.models.DNALibrary.index_format_choices,
        label="Index format",
        required=False,
    )

    min_num_read_groups = forms.IntegerField(
        label="Minimum number of read groups",
        min_value=0,
        required=False,
    )

    def clean_tagged_with(self):
        tags = self.cleaned_data['tagged_with']
        if tags:
            tags_list = [tag.strip() for tag in tags.split(",")]
            results = tantalus.models.SequenceDataset.objects.all()
            for tag in tags_list:
                if not results.filter(tags__name=tag).exists():
                    raise forms.ValidationError("Filter for the following tags together resulted in 0 results: {}".format(
                        ", ".join(tags_list)
                    ))
        return tags

    def clean_sample(self):
        sample = self.cleaned_data['sample']
        if sample:
            no_match_samples = []
            for samp in sample.split():
                if not tantalus.models.SequenceDataset.objects.filter(sample__sample_id=samp).exists():
                    no_match_samples.append(samp)
            if no_match_samples != []:
                raise forms.ValidationError("Filter for the following sample resulted in 0 results: {}".format(
                    ", ".join(no_match_samples)
                ))
        return sample

    def clean_library(self):
        library = self.cleaned_data['library']
        if library:
            no_match_list = []
            for lib in library.split():
                if not tantalus.models.SequenceDataset.objects.filter(library__library_id=lib).exists():
                    no_match_list.append(lib)

            if no_match_list:
                raise forms.ValidationError("Filter for the following library resulted in 0 results: {}".format(
                    ", ".join(no_match_list)
                ))
        return library

    def clean_flowcell_id_and_lane(self):
        flowcell_and_lane_number_input = self.cleaned_data['flowcell_id_and_lane']
        if flowcell_and_lane_number_input:
            no_match_list = []
            for flowcell_lane in flowcell_and_lane_number_input.split():
                if "_" in flowcell_lane:
                    # parse out flowcell ID and lane number, assumed to be separated by an underscore
                    flowcell, lane_number = flowcell_lane.split("_", 1)
                    if not tantalus.models.SequenceDataset.objects.filter(
                        sequence_lanes__flowcell_id=flowcell,sequence_lanes__lane_number=lane_number).exists():
                        no_match_list.append(flowcell_lane)
                else:
                    # no lane number included
                    if not tantalus.models.SequenceDataset.objects.filter(sequence_lanes__flowcell_id=flowcell_lane).exists():
                        no_match_list.append(flowcell_lane)
            if no_match_list:
                raise forms.ValidationError("Filter for the following flowcell lane resulted in 0 results: {}".format(
                    ", ".join(no_match_list)
                ))
        return flowcell_and_lane_number_input

    def clean_sequencing_library_id(self):
        sequencing_library_id_field = self.cleaned_data['sequencing_library_id']
        if sequencing_library_id_field:
            no_match_list = []
            for sequencing_library in sequencing_library_id_field.split():
                if not tantalus.models.SequenceDataset.objects.filter(library__library_id=sequencing_library).exists():
                    no_match_list.append(sequencing_library)
            if no_match_list:
                raise forms.ValidationError("Filter for the following sequencing library resulted in 0 results: {}".format(
                    no_match_list
                ))
        return sequencing_library_id_field

    def clean(self):
        cleaned_data = super(DatasetSearchForm, self).clean()
        results = self.get_dataset_search_results(clean=False, **cleaned_data)

        if len(results) == 0:
            raise forms.ValidationError(
                "Found zero datasets."
            )

    def get_dataset_search_results(self, clean=True, exclude=None, tagged_with=None, library=None, sample=None, dataset_type=None,storages=None,
                                   compression_schemes=None,flowcell_id_and_lane=None, sequencing_center=None,
                                   sequencing_instrument=None, sequencing_library_id=None, library_type=None,
                                   index_format=None, min_num_read_groups=None):
        """
        Performs the filter search with the given fields. The "clean" flag is used to indicate whether the cleaned data
        should be used or not.
            - This method gets called in the cleaning method - where the data is not yet guaranteed to be clean,
            and also outside, where the data can be trusted to be clean

        :param tags: list of tag strings separated by commas
        :param library: Library id. Eg. A90652A
        :param sample: Sample id. Eg. SA928
        :param clean: Flag indicating whether or not the data has been cleaned yet
        :return:
        """

        if clean:
            tagged_with = self.cleaned_data['tagged_with']
            exclude = self.cleaned_data['exclude']
            library = self.cleaned_data['library']
            sample = self.cleaned_data['sample']
            dataset_type = self.cleaned_data['dataset_type']
            storages = self.cleaned_data['storages']
            compression_schemes = self.cleaned_data['compression_schemes']
            flowcell_id_and_lane = self.cleaned_data['flowcell_id_and_lane']
            sequencing_center = self.cleaned_data['sequencing_center']
            sequencing_instrument = self.cleaned_data['sequencing_instrument']
            sequencing_library_id = self.cleaned_data['sequencing_library_id']
            library_type = self.cleaned_data['library_type']
            index_format = self.cleaned_data['index_format']
            min_num_read_groups = self.cleaned_data['min_num_read_groups']


        results = tantalus.models.SequenceDataset.objects.all()

        # TODO: add prefetch related

        if tagged_with:
            tags_list = [tag.strip() for tag in tagged_with.split(",")]
            exclude_list = [tag.strip() for tag in exclude.split(",")]
            for tag in tags_list:
                results = results.filter(tags__name=tag).exclude(tags__name__in=exclude_list)

        if sample:
            results = results.filter(sample__sample_id__in=sample.split())

        if dataset_type:
            results = results.filter(dataset_type__in=dataset_type)

        if storages:
            results = results.filter(file_resources__fileinstance__storage_id__in=storages)

        if compression_schemes:
            results = results.filter(file_resources__compression__in=compression_schemes)

        if library:
            results = results.filter(library__library_id__in=library.split())

        if sequencing_center:
            results = results.filter(sequence_lanes__sequencing_centre=sequencing_center)

        if sequencing_instrument:
            results = results.filter(sequence_lanes__sequencing_instrument=sequencing_instrument)

        if sequencing_library_id:
            results = results.filter(library__library_id__in=sequencing_library_id.split())

        if library_type:
            results = results.filter(library__library_type=library_type)

        if index_format:
            results = results.filter(library__index_format=index_format)

        if min_num_read_groups is not None:
            results = results.annotate(sequence_lanes__lane_number=Count('sequence_lanes__lane_number')).filter(sequence_lanes__lane_number__gte=min_num_read_groups)

        if flowcell_id_and_lane:
            query = Q()
            for flowcell_lane in flowcell_id_and_lane.split():
                if "_" in flowcell_lane:
                    # parse out flowcell ID and lane number, assumed to be separated by an underscore
                    flowcell, lane_number = flowcell_lane.split("_", 1)
                    q = Q(sequence_lanes__flowcell_id=flowcell, sequence_lanes__lane_number=lane_number)
                else:
                    q = Q(sequence_lanes__flowcell_id=flowcell_lane)
                query = query | q
            results = results.filter(query)

        results = results.distinct()

        return list(results.values_list('id', flat=True))


class AddDatasetToTagForm(forms.Form):

    tag = forms.CharField(max_length=40)

    def clean_tag(self):
        tag_name = self.cleaned_data.get("tag", False).strip()
        tag, created = tantalus.models.Tag.objects.get_or_create(name=tag_name)
        return tag


class DatasetTagForm(forms.Form):
    tag_name = forms.CharField(max_length=500)
    models_to_tag = None

    # use __init__ to populate models to tag field
    def __init__(self, *args, **kwargs):
        datasets = kwargs.pop('datasets', None)
        super(DatasetTagForm, self).__init__(*args, **kwargs)

        if datasets:
            self.models_to_tag = tantalus.models.SequenceDataset.objects.filter(pk__in=datasets)
        else:
            self.models_to_tag = tantalus.models.SequenceDataset.objects.all()

    def add_dataset_tags(self):
        tag_name = self.cleaned_data['tag_name']
        tag, created = tantalus.models.Tag.objects.get_or_create(name=tag_name)
        tag.sequencedataset_set.clear()
        tag.sequencedataset_set.add(*self.models_to_tag)

